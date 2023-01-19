# working_with_excel.py

# loading a workbook and sheet
from openpyxl import Workbook, load_workbook

# #create instance of load_workbook with pathname to excel file
# wb = load_workbook("/Users/dannelson/Desktop/misc_templates/grades.xlsx")

# #Select the active sheet
# sheet = wb.active

# print(sheet)

# # accessing a cell value
# print(sheet['A1'].value)
# print(sheet['A2'].value)
# print(sheet['B3'].value)

# # changing a cell value
# sheet['A2'] = 'Test'

# # saving workbook
# wb.save('/Users/dannelson/Desktop/misc_templates/grades.xlsx')

# # creating a new sheet
# wb.create_sheet("Class B")

# # return list of sheets
# print(wb.sheetnames)

# # create new workbook
# new_wb = Workbook()

# #select active sheet
# ws = new_wb.active

# #rename active sheet
# ws.title  = "Test"

# #add data to sheet
# ws.append(["This","Is","A","Test"])

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")


# #add more rows to sheet
# ws.append(["This","Is","Another","Test"])
# ws.append(["And","Yet","Another","Test"])
# ws.append(["End"])

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")


# #add more rows to sheet
# ws.append(["This","Is","Another","Test"])
# ws.append(["And","Yet","Another","Test"])
# ws.append(["End"])

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")

#  # Accessing multiple cell values
#  #import get_column_letter
# from openpyxl.utils import get_column_letter

# #access multiple cell values
# for row in range(1,5):
#     for col in range(1,5):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)

# #access multiple cell values 
# for row in range(1,ws.max_row + 1):
#     for col in range(1,ws.max_column + 1):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)

# #modify multiple cell values 
# for row in range(1,ws.max_row + 1):
#     for col in range(1,ws.max_column + 1):
#         char = get_column_letter(col)
#         ws[char + str(row)] = char + str(row)
        
# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")

# #insert new row
# ws.insert_rows(1)

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")

# #delete a row
# ws.delete_rows(1)

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")


# #merge cells
# ws.merge_cells("A1:A2")

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")


# #unmerge cells
# ws.unmerge_cells("A1:A2")

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")

# #insert columns
# ws.insert_cols(2)

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")

# #delete columns
# ws.delete_cols(1,2)

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")

# #move range
# ws.move_range("A1:B1",rows=4,cols=1)

# #save
# new_wb.save("/Users/dannelson/Desktop/misc_templates/test.xlsx")


#Sales data Python dictionary
sales_data = {
           "XpressWear Pants":{"New York":5000,"Toronto":7000,"London":6000,"Hong Kong":10000},
           "Swish Wallet ":{"New York":8000,"Toronto":3000,"London":5000,"Hong Kong":9000},
           "ONEset Shaving Kit":{"New York":6000,"Toronto":9000,"London":4000,"Hong Kong":6000},
           "Rhino Phone Case":{"New York":2000,"Toronto":5000,"London":4000,"Hong Kong":7000}
            }
#create new workbook
sales_wb = Workbook()

#select active sheet
ws = sales_wb.active

#change name of sheet
ws.title = "Sales"

#create the column headers with dictionary keys
column_names = ["Product Name"] + list(sales_data["XpressWear Pants"].keys())

#append column names to sheet
ws.append(column_names)

#loop through dictionary and append values to sheet
for product in sales_data:
    sales = list(sales_data[product].values())
    ws.append([product] + sales)

#save
sales_wb.save("/Users/dannelson/Desktop/misc_templates/sales_data.xlsx")

#aggregating a range of cells
ws['B6'] = '=AVERAGE(B2:B5)'

#save
sales_wb.save("/Users/dannelson/Desktop/misc_templates/sales_data.xlsx")

def get_column_letter(num):
    letters = {1: "A", 2: "B", 3: "C", 4: 'D', 5: "E", 6: "F"}
    col_name = letters[num]
    return col_name

#add title of aggregation
ws['A'+str(ws.max_row+1)] = "Total Sales"

#aggregate column values
for col in range(2, len(sales_data["XpressWear Pants"]) + 2):
    char = get_column_letter(col) # this is not defined-must be a function not yet defined
    ws[char + '6'] = f"=SUM({char+'2'}:{char+'5'})"
        
#save
sales_wb.save("/Users/dannelson/Desktop/misc_templates/sales_data.xlsx")


# styling the font
from openpyxl.styles import Font

#apply a bold font
for col in range(1,ws.max_column+1):
    ws[get_column_letter(col) + '1'].font = Font('Arial', bold=True, size=13, color='00000080')
    
#save
sales_wb.save("/Users/dannelson/Desktop/misc_templates/sales_data.xlsx")


#import barchart module
from openpyxl.chart import BarChart, Reference

#create bar chart object
barchart = BarChart()

#select data range
data = Reference(ws, min_col=ws.min_column+1, max_col=ws.max_column, min_row=ws.min_row, max_row=ws.max_row-1)

#specify product names as categories
categories = Reference(ws, min_col=ws.min_column, max_col=ws.min_column, min_row=ws.min_row+1, max_row=ws.max_row-1)

#add data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#location of chart
ws.add_chart(barchart, "G1")

#add chart title
barchart.title = 'Product Sales by City'

#choose the chart style
barchart.style = 2

#save
sales_wb.save("/Users/dannelson/Desktop/misc_templates/sales_data.xlsx")

# Read the official documentation at:
official_docs = 'https://openpyxl.readthedocs.io/en/stable/index.html'
